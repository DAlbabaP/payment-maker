"""
Модуль генерации документов Excel
"""

import logging
from pathlib import Path
from typing import List, Optional
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from paymentmaker.core.models import InvoiceData, ActData, CompanyDetails, TransportService
from paymentmaker.templates.invoice import InvoiceTemplate
from paymentmaker.templates.act import ActTemplate
from paymentmaker.utils.constants import DEFAULT_SIGNATURES

logger = logging.getLogger(__name__)


class DocumentGenerator:
    """Генератор документов Excel"""
    
    def __init__(self, company: CompanyDetails, customer: CompanyDetails):
        self.company = company
        self.customer = customer
        self.invoice_template = InvoiceTemplate(company, customer)
        self.act_template = ActTemplate(company, customer)
    
    def generate_documents(self, 
                         invoice_data: InvoiceData,
                         output_path: str) -> bool:
        """
        Генерирует счет и акт в одном файле Excel
        
        Args:
            invoice_data: Данные для счета/акта
            output_path: Путь для сохранения
            
        Returns:
            True если успешно, False в случае ошибки
        """
        try:
            # Создаем новую книгу
            wb = openpyxl.Workbook()
            
            # Удаляем стандартный лист
            wb.remove(wb.active)
            
            # Создаем листы для счета и акта
            invoice_sheet = wb.create_sheet("Счет")
            act_sheet = wb.create_sheet("Акт")
            
            # Генерируем счет
            self.invoice_template.create(invoice_sheet, invoice_data)
            
            # Создаем данные для акта на основе счета
            act_data = ActData(
                number=invoice_data.number,
                date=invoice_data.date,
                customer=invoice_data.customer,
                customer_details=invoice_data.customer_details,
                services=invoice_data.services
            )
            
            # Генерируем акт
            self.act_template.create(act_sheet, act_data)
            
            # Сохраняем файл
            output_path = Path(output_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(str(output_path))
            
            logger.info(f"Документы успешно сохранены: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Ошибка при генерации документов: {e}")
            return False
    
    def update_template(self,
                       template_path: str,
                       invoice_data: InvoiceData,
                       output_path: str) -> bool:
        """
        Обновляет существующий шаблон данными
        
        Args:
            template_path: Путь к файлу шаблона
            invoice_data: Данные для заполнения
            output_path: Путь для сохранения результата
            
        Returns:
            True если успешно, False в случае ошибки
        """
        try:
            # Загружаем существующий шаблон
            wb = openpyxl.load_workbook(template_path)
            
            # Проверяем наличие необходимых листов
            if "Счет" not in wb.sheetnames or "Акт" not in wb.sheetnames:
                logger.error("Шаблон не содержит необходимых листов")
                return False
            
            # Заполняем счет
            invoice_sheet = wb["Счет"]
            self._fill_invoice_data(invoice_sheet, invoice_data)
            
            # Заполняем акт
            act_sheet = wb["Акт"]
            act_data = ActData(
                number=invoice_data.number,
                date=invoice_data.date,
                customer=invoice_data.customer,
                customer_details=invoice_data.customer_details,
                services=invoice_data.services
            )
            self._fill_act_data(act_sheet, act_data)
            
            # Сохраняем результат
            output_path = Path(output_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(str(output_path))
            
            logger.info(f"Шаблон успешно заполнен: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Ошибка при заполнении шаблона: {e}")
            return False
    
    def _fill_invoice_data(self, sheet: Worksheet, invoice_data: InvoiceData):
        """Заполняет данные счета на листе"""
        # Обновляем заголовок
        sheet['A10'].value = f"Счет на оплату № {invoice_data.number} от {invoice_data.date_str}"
        
        # Обновляем данные клиента
        if invoice_data.customer:
            sheet['C14'].value = invoice_data.customer
        
        # Заполняем таблицу услуг
        self._fill_services_table(sheet, invoice_data.services, start_row=17)
        
        # Обновляем итоги
        self._update_totals(sheet, invoice_data, start_row=17 + len(invoice_data.services))
    
    def _fill_act_data(self, sheet: Worksheet, act_data: ActData):
        """Заполняет данные акта на листе"""
        # Обновляем заголовок
        sheet['A1'].value = f"Акт № {act_data.number} от {act_data.date_str}"
        
        # Обновляем данные клиента
        if act_data.customer:
            sheet['C6'].value = act_data.customer
        
        # Обновляем основание
        sheet['C8'].value = f"По счету № {act_data.number} от {act_data.date_str}"
        
        # Заполняем таблицу услуг
        self._fill_services_table(sheet, act_data.services, start_row=11)
        
        # Обновляем итоги
        self._update_totals(sheet, act_data, start_row=11 + len(act_data.services))
    
    def _fill_services_table(self, sheet: Worksheet, services: List[TransportService], start_row: int):
        """Заполняет таблицу услуг"""
        from paymentmaker.templates.styles import DocumentStyles
        
        border = DocumentStyles.get_thin_border()
        normal_font = DocumentStyles.get_normal_font()
        small_font = DocumentStyles.get_small_font()
        
        for i, service in enumerate(services):
            row = start_row + i
            
            # Номер
            sheet[f'A{row}'].value = i + 1
            sheet[f'A{row}'].border = border
            sheet[f'A{row}'].alignment = DocumentStyles.get_center_alignment()
            sheet[f'A{row}'].font = normal_font
            
            # Описание
            sheet[f'B{row}'].value = service.description
            sheet[f'B{row}'].border = border
            sheet[f'B{row}'].alignment = DocumentStyles.get_wrap_text_alignment()
            sheet[f'B{row}'].font = small_font
            sheet.row_dimensions[row].height = 30
            
            # Количество
            sheet[f'C{row}'].value = service.quantity
            sheet[f'C{row}'].border = border
            sheet[f'C{row}'].alignment = DocumentStyles.get_center_alignment()
            sheet[f'C{row}'].font = normal_font
            
            # Единица измерения
            sheet[f'D{row}'].value = service.unit
            sheet[f'D{row}'].border = border
            sheet[f'D{row}'].alignment = DocumentStyles.get_center_alignment()
            sheet[f'D{row}'].font = normal_font
            
            # Цена
            sheet[f'E{row}'].value = float(service.price)
            sheet[f'E{row}'].border = border
            sheet[f'E{row}'].alignment = DocumentStyles.get_center_alignment()
            sheet[f'E{row}'].font = normal_font
            sheet[f'E{row}'].number_format = '#,##0.00'
            
            # Сумма
            sheet[f'F{row}'].value = float(service.amount)
            sheet[f'F{row}'].border = border
            sheet[f'F{row}'].alignment = DocumentStyles.get_center_alignment()
            sheet[f'F{row}'].font = normal_font
            sheet[f'F{row}'].number_format = '#,##0.00'
    
    def _update_totals(self, sheet: Worksheet, data: InvoiceData, start_row: int):
        """Обновляет итоговые суммы"""
        from paymentmaker.templates.styles import DocumentStyles
        
        border = DocumentStyles.get_thin_border()
        header_font = DocumentStyles.get_header_font()
        
        total_amount = float(data.total_amount)
        
        # Итого
        sheet.merge_cells(f'A{start_row}:E{start_row}')
        sheet[f'A{start_row}'].value = "Итого:"
        sheet[f'A{start_row}'].alignment = DocumentStyles.get_right_alignment()
        sheet[f'A{start_row}'].font = header_font
        sheet[f'F{start_row}'].value = total_amount
        sheet[f'F{start_row}'].border = border
        sheet[f'F{start_row}'].font = header_font
        sheet[f'F{start_row}'].number_format = '#,##0.00'
        
        # НДС
        sheet.merge_cells(f'A{start_row+1}:E{start_row+1}')
        sheet[f'A{start_row+1}'].value = "Без налога (НДС)"
        sheet[f'A{start_row+1}'].alignment = DocumentStyles.get_right_alignment()
        sheet[f'A{start_row+1}'].font = header_font
        sheet[f'F{start_row+1}'].value = '-'
        sheet[f'F{start_row+1}'].border = border
        sheet[f'F{start_row+1}'].font = header_font
        
        # Сумма прописью
        amount_text = self._amount_to_words(total_amount)
        
        sheet.merge_cells(f'A{start_row+2}:F{start_row+2}')
        sheet[f'A{start_row+2}'].value = f"Всего наименований {data.services_count}, на сумму {total_amount:,.2f} руб."
        sheet[f'A{start_row+2}'].font = header_font
        
        sheet.merge_cells(f'A{start_row+3}:F{start_row+3}')
        sheet[f'A{start_row+3}'].value = amount_text
        sheet[f'A{start_row+3}'].font = header_font
    
    def _amount_to_words(self, amount: float) -> str:
        """Преобразует сумму в текст"""
        try:
            from num2words import num2words
            amount_str = num2words(int(amount), lang='ru')
            amount_str = amount_str[0].upper() + amount_str[1:]
            return f"{amount_str} рублей 00 копеек. Без НДС."
        except Exception:
            return f"{amount:,.2f} рублей 00 копеек. Без НДС."