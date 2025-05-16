"""
Шаблон счета
"""

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.page import PageMargins

from paymentmaker.core.models import InvoiceData, CompanyDetails
from paymentmaker.templates.styles import DocumentStyles


class InvoiceTemplate:
    """Шаблон для создания счета"""
    
    def __init__(self, company: CompanyDetails, customer: CompanyDetails):
        self.company = company
        self.customer = customer
    
    def create(self, sheet: Worksheet, invoice_data: InvoiceData):
        """Создает счет на листе Excel"""
        # Настройка страницы
        self._setup_page(sheet)
        
        # Создание структуры счета
        self._create_header_section(sheet)
        self._create_bank_section(sheet)
        self._create_invoice_title(sheet, invoice_data)
        self._create_parties_section(sheet, invoice_data)
        self._create_table_header(sheet)
        self._fill_services_table(sheet, invoice_data)
        self._create_totals_section(sheet, invoice_data)
        self._create_signatures_section(sheet)
        
        # Установка области печати
        row_count = 17 + len(invoice_data.services) + 10
        sheet.print_area = f'A1:F{row_count}'
    
    def _setup_page(self, sheet: Worksheet):
        """Настройка параметров страницы"""
        # Ширина колонок
        sheet.column_dimensions['A'].width = 3
        sheet.column_dimensions['B'].width = 55
        sheet.column_dimensions['C'].width = 8
        sheet.column_dimensions['D'].width = 6
        sheet.column_dimensions['E'].width = 15
        sheet.column_dimensions['F'].width = 20
        
        # Настройки печати
        sheet.page_setup.orientation = 'portrait'
        sheet.page_setup.paperSize = 9  # A4
        sheet.page_setup.fitToPage = True
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 1
        
        # Поля
        sheet.page_margins = PageMargins(
            left=0.2, right=0.2, top=0.3, bottom=0.3,
            header=0.2, footer=0.2
        )
    
    def _create_header_section(self, sheet: Worksheet):
        """Создает верхнюю часть счета с предупреждением"""
        # Объединяем ячейки для предупреждения
        sheet.merge_cells('A1:F2')
        cell = sheet['A1']
        cell.value = (
            "Внимание! Оплата данного счета означает согласие с условиями поставки товара. "
            "Уведомление об оплате обязательно, в противном случае не гарантируется наличие "
            "товара на складе. Товар отпускается по факту прихода денег на р/с Поставщика, "
            "самовывозом, при наличии доверенности и паспорта."
        )
        cell.alignment = DocumentStyles.get_wrap_text_alignment()
        cell.font = DocumentStyles.get_small_font()
        
        # Настройка высоты строк
        sheet.row_dimensions[1].height = 20
        sheet.row_dimensions[2].height = 20
    
    def _create_bank_section(self, sheet: Worksheet):
        """Создает секцию с банковскими реквизитами"""
        # Название банка
        sheet.merge_cells('A3:D4')
        sheet['A3'] = self.company.bank_name
        sheet['A3'].font = DocumentStyles.get_normal_font()
        
        # БИК
        sheet['E3'] = "БИК"
        sheet['F3'] = self.company.bank_bik
        sheet['F3'].font = DocumentStyles.get_normal_font()
        
        # Счет банка
        sheet.merge_cells('E4:E5')
        sheet['E4'] = "Сч. №"
        sheet.merge_cells('F4:F5')
        sheet['F4'] = self.company.bank_account
        sheet['F4'].font = DocumentStyles.get_small_font()
        
        # Банк получателя
        sheet.merge_cells('A5:D5')
        sheet['A5'] = "Банк получателя"
        sheet['A5'].font = DocumentStyles.get_normal_font()
        
        # ИНН получателя
        sheet['A6'] = f"ИНН {self.company.inn}"
        sheet['A6'].font = DocumentStyles.get_normal_font()
        
        # КПП
        sheet.merge_cells('C6:D6')
        sheet['C6'] = "КПП"
        sheet['C6'].font = DocumentStyles.get_normal_font()
        
        # Счет получателя
        sheet.merge_cells('E6:E8')
        sheet['E6'] = "Сч. №"
        sheet.merge_cells('F6:F8')
        sheet['F6'] = self.company.company_account
        sheet['F6'].font = DocumentStyles.get_small_font()
        
        # Название получателя
        sheet.merge_cells('A7:D7')
        sheet['A7'] = self.company.name
        sheet['A7'].font = DocumentStyles.get_normal_font()
        
        # Получатель
        sheet.merge_cells('A8:D8')
        sheet['A8'] = "Получатель"
        sheet['A8'].font = DocumentStyles.get_normal_font()
        
        # Добавляем границы
        border = DocumentStyles.get_thin_border()
        for row in range(3, 9):
            for col in range(1, 7):
                sheet.cell(row=row, column=col).border = border
    
    def _create_invoice_title(self, sheet: Worksheet, invoice_data: InvoiceData):
        """Создает заголовок счета"""
        sheet.merge_cells('A10:F10')
        cell = sheet['A10']
        cell.value = f"Счет на оплату № {invoice_data.number} от {invoice_data.date_str}"
        cell.font = DocumentStyles.get_title_font()
        cell.alignment = DocumentStyles.get_center_alignment()
        
        # Толстая нижняя граница
        thick_border = DocumentStyles.get_thick_border()
        for col in range(1, 7):
            sheet.cell(row=10, column=col).border = thick_border
    
    def _create_parties_section(self, sheet: Worksheet, invoice_data: InvoiceData):
        """Создает секцию с информацией о поставщике и покупателе"""
        # Поставщик
        sheet.merge_cells('A12:B12')
        sheet['A12'] = "Поставщик:"
        sheet['A12'].font = DocumentStyles.get_normal_font()
        
        sheet.merge_cells('C12:F12')
        sheet['C12'] = self.company.full_details
        sheet['C12'].font = DocumentStyles.get_small_font()
        sheet['C12'].alignment = DocumentStyles.get_wrap_text_alignment()
        sheet.row_dimensions[12].height = 25
        
        # Покупатель
        sheet.merge_cells('A14:B14')
        sheet['A14'] = "Покупатель:"
        sheet['A14'].font = DocumentStyles.get_normal_font()
        
        sheet.merge_cells('C14:F14')
        sheet['C14'] = invoice_data.customer or self.customer.full_details
        sheet['C14'].font = DocumentStyles.get_small_font()
        sheet['C14'].alignment = DocumentStyles.get_wrap_text_alignment()
        sheet.row_dimensions[14].height = 30
    
    def _create_table_header(self, sheet: Worksheet):
        """Создает заголовок таблицы"""
        headers = ["№", "Товары (работы, услуги)", "Кол-во", "Ед.", "Цена", "Сумма"]
        
        for col_idx, header in enumerate(headers):
            cell = sheet.cell(row=16, column=col_idx + 1)
            cell.value = header
            cell.font = DocumentStyles.get_header_font()
            cell.alignment = DocumentStyles.get_center_alignment()
            cell.border = DocumentStyles.get_thin_border()
            cell.fill = DocumentStyles.get_header_fill()
    
    def _fill_services_table(self, sheet: Worksheet, invoice_data: InvoiceData):
        """Заполняет таблицу услуг"""
        border = DocumentStyles.get_thin_border()
        normal_font = DocumentStyles.get_normal_font()
        small_font = DocumentStyles.get_small_font()
        
        start_row = 17
        
        for i, service in enumerate(invoice_data.services):
            row = start_row + i
            
            # Номер
            sheet[f'A{row}'] = i + 1
            sheet[f'A{row}'].border = border
            sheet[f'A{row}'].alignment = DocumentStyles.get_center_alignment()
            sheet[f'A{row}'].font = normal_font
            
            # Описание
            sheet[f'B{row}'] = service.description
            sheet[f'B{row}'].border = border
            sheet[f'B{row}'].alignment = DocumentStyles.get_wrap_text_alignment()
            sheet[f'B{row}'].font = small_font
            sheet.row_dimensions[row].height = 30
            
            # Количество
            sheet[f'C{row}'] = service.quantity
            sheet[f'C{row}'].border = border
            sheet[f'C{row}'].alignment = DocumentStyles.get_center_alignment()
            sheet[f'C{row}'].font = normal_font
            
            # Единица
            sheet[f'D{row}'] = service.unit
            sheet[f'D{row}'].border = border
            sheet[f'D{row}'].alignment = DocumentStyles.get_center_alignment()
            sheet[f'D{row}'].font = normal_font
            
            # Цена
            sheet[f'E{row}'] = float(service.price)
            sheet[f'E{row}'].border = border
            sheet[f'E{row}'].alignment = DocumentStyles.get_center_alignment()
            sheet[f'E{row}'].font = normal_font
            sheet[f'E{row}'].number_format = '#,##0.00'
            
            # Сумма
            sheet[f'F{row}'] = float(service.amount)
            sheet[f'F{row}'].border = border
            sheet[f'F{row}'].alignment = DocumentStyles.get_center_alignment()
            sheet[f'F{row}'].font = normal_font
            sheet[f'F{row}'].number_format = '#,##0.00'
    
    def _create_totals_section(self, sheet: Worksheet, invoice_data: InvoiceData):
        """Создает секцию с итогами"""
        start_row = 17 + len(invoice_data.services)
        border = DocumentStyles.get_thin_border()
        header_font = DocumentStyles.get_header_font()
        
        # Итого
        sheet.merge_cells(f'A{start_row}:E{start_row}')
        sheet[f'A{start_row}'] = "Итого:"
        sheet[f'A{start_row}'].alignment = DocumentStyles.get_right_alignment()
        sheet[f'A{start_row}'].font = header_font
        sheet[f'F{start_row}'] = float(invoice_data.total_amount)
        sheet[f'F{start_row}'].border = border
        sheet[f'F{start_row}'].font = header_font
        sheet[f'F{start_row}'].number_format = '#,##0.00'
        
        # Без НДС
        sheet.merge_cells(f'A{start_row+1}:E{start_row+1}')
        sheet[f'A{start_row+1}'] = "Без налога (НДС)"
        sheet[f'A{start_row+1}'].alignment = DocumentStyles.get_right_alignment()
        sheet[f'A{start_row+1}'].font = header_font
        sheet[f'F{start_row+1}'] = '-'
        sheet[f'F{start_row+1}'].border = border
        sheet[f'F{start_row+1}'].font = header_font
        
        # Сумма прописью
        amount_text = self._amount_to_words(float(invoice_data.total_amount))
        
        sheet.merge_cells(f'A{start_row+2}:F{start_row+2}')
        sheet[f'A{start_row+2}'] = f"Всего наименований {invoice_data.services_count}, на сумму {invoice_data.total_amount:,.2f} руб."
        sheet[f'A{start_row+2}'].font = header_font
        
        sheet.merge_cells(f'A{start_row+3}:F{start_row+3}')
        sheet[f'A{start_row+3}'] = amount_text
        sheet[f'A{start_row+3}'].font = header_font
    
    def _create_signatures_section(self, sheet: Worksheet):
        """Создает секцию подписей"""
        start_row = len(sheet['B']) + 5
        header_font = DocumentStyles.get_header_font()
        
        # Руководитель
        sheet[f'A{start_row}'] = "Руководитель"
        sheet[f'A{start_row}'].font = header_font
        
        # ФИО руководителя
        try:
            sheet.unmerge_cells(f'C{start_row}:D{start_row}')
        except:
            pass
        sheet.merge_cells(f'C{start_row}:D{start_row}')
        sheet[f'C{start_row}'] = "_______________"
        sheet[f'C{start_row}'].font = header_font
        
        # Бухгалтер
        sheet[f'E{start_row}'] = "Бухгалтер"
        sheet[f'E{start_row}'].font = header_font
        
        # ФИО бухгалтера
        try:
            sheet.unmerge_cells(f'F{start_row}:G{start_row}')
        except:
            pass
        sheet.merge_cells(f'F{start_row}:G{start_row}')
        sheet[f'F{start_row}'] = "_______________"
        sheet[f'F{start_row}'].font = header_font
    
    def _amount_to_words(self, amount: float) -> str:
        """Преобразует сумму в слова"""
        try:
            from num2words import num2words
            amount_str = num2words(int(amount), lang='ru')
            amount_str = amount_str[0].upper() + amount_str[1:]
            return f"{amount_str} рублей 00 копеек. Без НДС."
        except:
            return f"{amount:,.2f} рублей 00 копеек. Без НДС."