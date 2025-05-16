"""
Шаблон акта выполненных работ
"""

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.page import PageMargins

from paymentmaker.core.models import ActData, CompanyDetails
from paymentmaker.templates.styles import DocumentStyles


class ActTemplate:
    """Шаблон для создания акта выполненных работ"""
    
    def __init__(self, company: CompanyDetails, customer: CompanyDetails):
        self.company = company
        self.customer = customer
    
    def create(self, sheet: Worksheet, act_data: ActData):
        """Создает акт на листе Excel"""
        # Настройка страницы
        self._setup_page(sheet)
        
        # Создание структуры акта
        self._create_act_title(sheet, act_data)
        self._create_parties_section(sheet, act_data)
        self._create_basis_section(sheet, act_data)
        self._create_table_header(sheet)
        self._fill_services_table(sheet, act_data)
        self._create_totals_section(sheet, act_data)
        self._create_conclusion_section(sheet)
        self._create_signatures_section(sheet)
        
        # Установка области печати
        row_count = 11 + len(act_data.services) + 15
        sheet.print_area = f'A1:F{row_count}'
    
    def _setup_page(self, sheet: Worksheet):
        """Настройка параметров страницы"""
        # Ширина колонок
        sheet.column_dimensions['A'].width = 3
        sheet.column_dimensions['B'].width = 55
        sheet.column_dimensions['C'].width = 8
        sheet.column_dimensions['D'].width = 6
        sheet.column_dimensions['E'].width = 15
        sheet.column_dimensions['F'].width = 20
        
        # Настройки печати
        sheet.page_setup.orientation = 'portrait'
        sheet.page_setup.paperSize = 9  # A4
        sheet.page_setup.fitToPage = True
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 1
        
        # Поля
        sheet.page_margins = PageMargins(
            left=0.2, right=0.2, top=0.3, bottom=0.3,
            header=0.2, footer=0.2
        )
    
    def _create_act_title(self, sheet: Worksheet, act_data: ActData):
        """Создает заголовок акта"""
        sheet.merge_cells('A1:F1')
        cell = sheet['A1']
        cell.value = f"Акт № {act_data.number} от {act_data.date_str}"
        cell.font = DocumentStyles.get_title_font()
        cell.alignment = DocumentStyles.get_center_alignment()
        
        # Толстая нижняя граница
        thick_border = DocumentStyles.get_thick_border()
        for col in range(1, 7):
            sheet.cell(row=1, column=col).border = thick_border
    
    def _create_parties_section(self, sheet: Worksheet, act_data: ActData):
        """Создает секцию с информацией об исполнителе и заказчике"""
        # Исполнитель
        sheet.merge_cells('A3:B3')
        sheet['A3'] = "Исполнитель:"
        sheet['A3'].font = DocumentStyles.get_normal_font()
        
        sheet.merge_cells('C3:F3')
        sheet['C3'] = self.company.full_details
        sheet['C3'].font = DocumentStyles.get_small_font()
        sheet['C3'].alignment = DocumentStyles.get_wrap_text_alignment()
        sheet.row_dimensions[3].height = 35
        
        # Реквизиты исполнителя
        sheet.merge_cells('C4:F4')
        sheet['C4'] = self.company.bank_details
        sheet['C4'].font = DocumentStyles.get_small_font()
        sheet['C4'].alignment = DocumentStyles.get_wrap_text_alignment()
        sheet.row_dimensions[4].height = 35
        
        # Заказчик
        sheet.merge_cells('A6:B6')
        sheet['A6'] = "Заказчик:"
        sheet['A6'].font = DocumentStyles.get_normal_font()
        
        sheet.merge_cells('C6:F6')
        sheet['C6'] = act_data.customer or self.customer.full_details
        sheet['C6'].font = DocumentStyles.get_small_font()
        sheet['C6'].alignment = DocumentStyles.get_wrap_text_alignment()
        sheet.row_dimensions[6].height = 35
        
        # Реквизиты заказчика
        sheet.merge_cells('C7:F7')
        sheet['C7'] = act_data.customer_details or self.customer.bank_details
        sheet['C7'].font = DocumentStyles.get_small_font()
        sheet['C7'].alignment = DocumentStyles.get_wrap_text_alignment()
        sheet.row_dimensions[7].height = 25
    
    def _create_basis_section(self, sheet: Worksheet, act_data: ActData):
        """Создает секцию с основанием"""
        sheet.merge_cells('A8:B8')
        sheet['A8'] = "Основание:"
        sheet['A8'].font = DocumentStyles.get_normal_font()
        
        sheet.merge_cells('C8:F8')
        sheet['C8'] = f"По счету № {act_data.number} от {act_data.date_str}"
        sheet['C8'].font = DocumentStyles.get_normal_font()
    
    def _create_table_header(self, sheet: Worksheet):
        """Создает заголовок таблицы"""
        headers = ["№", "Наименование работ, услуг", "Кол-во", "Ед.", "Цена", "Сумма"]
        
        for col_idx, header in enumerate(headers):
            cell = sheet.cell(row=10, column=col_idx + 1)
            cell.value = header
            cell.font = DocumentStyles.get_header_font()
            cell.alignment = DocumentStyles.get_center_alignment()
            cell.border = DocumentStyles.get_thin_border()
            cell.fill = DocumentStyles.get_header_fill()
    
    def _fill_services_table(self, sheet: Worksheet, act_data: ActData):
        """Заполняет таблицу услуг"""
        border = DocumentStyles.get_thin_border()
        normal_font = DocumentStyles.get_normal_font()
        small_font = DocumentStyles.get_small_font()
        
        start_row = 11
        
        for i, service in enumerate(act_data.services):
            row = start_row + i
            
            # Номер
            sheet[f'A{row}'] = i + 1
            sheet[f'A{row}'].border = border
            sheet[f'A{row}'].alignment = DocumentStyles.get_center_alignment()
            sheet[f'A{row}'].font = normal_font
            
            # Описание
            sheet[f'B{row}'] = service.description
            sheet[f'B{row}'].border = border
            sheet[f'B{row}'].alignment = DocumentStyles.get_wrap_text_alignment()
            sheet[f'B{row}'].font = small_font
            sheet.row_dimensions[row].height = 30
            
            # Количество
            sheet[f'C{row}'] = service.quantity
            sheet[f'C{row}'].border = border
            sheet[f'C{row}'].alignment = DocumentStyles.get_center_alignment()
            sheet[f'C{row}'].font = normal_font
            
            # Единица
            sheet[f'D{row}'] = service.unit
            sheet[f'D{row}'].border = border
            sheet[f'D{row}'].alignment = DocumentStyles.get_center_alignment()
            sheet[f'D{row}'].font = normal_font
            
            # Цена
            sheet[f'E{row}'] = float(service.price)
            sheet[f'E{row}'].border = border
            sheet[f'E{row}'].alignment = DocumentStyles.get_center_alignment()
            sheet[f'E{row}'].font = normal_font
            sheet[f'E{row}'].number_format = '#,##0.00'
            
            # Сумма
            sheet[f'F{row}'] = float(service.amount)
            sheet[f'F{row}'].border = border
            sheet[f'F{row}'].alignment = DocumentStyles.get_center_alignment()
            sheet[f'F{row}'].font = normal_font
            sheet[f'F{row}'].number_format = '#,##0.00'
    
    def _create_totals_section(self, sheet: Worksheet, act_data: ActData):
        """Создает секцию с итогами"""
        start_row = 11 + len(act_data.services)
        border = DocumentStyles.get_thin_border()
        header_font = DocumentStyles.get_header_font()
        
        # Итого
        sheet.merge_cells(f'A{start_row}:E{start_row}')
        sheet[f'A{start_row}'] = "Итого:"
        sheet[f'A{start_row}'].alignment = DocumentStyles.get_right_alignment()
        sheet[f'A{start_row}'].font = header_font
        sheet[f'F{start_row}'] = float(act_data.total_amount)
        sheet[f'F{start_row}'].border = border
        sheet[f'F{start_row}'].font = header_font
        sheet[f'F{start_row}'].number_format = '#,##0.00'
        
        # Без НДС
        sheet.merge_cells(f'A{start_row+1}:E{start_row+1}')
        sheet[f'A{start_row+1}'] = "Без налога (НДС)"
        sheet[f'A{start_row+1}'].alignment = DocumentStyles.get_right_alignment()
        sheet[f'A{start_row+1}'].font = header_font
        sheet[f'F{start_row+1}'] = '-'
        sheet[f'F{start_row+1}'].border = border
        sheet[f'F{start_row+1}'].font = header_font
        
        # Сумма прописью
        amount_text = self._amount_to_words(float(act_data.total_amount))
        
        sheet.merge_cells(f'A{start_row+3}:F{start_row+3}')
        sheet[f'A{start_row+3}'] = f"Всего оказано услуг {act_data.services_count}, на сумму {act_data.total_amount:,.2f} руб."
        sheet[f'A{start_row+3}'].font = header_font
        
        sheet.merge_cells(f'A{start_row+4}:F{start_row+4}')
        sheet[f'A{start_row+4}'] = amount_text
        sheet[f'A{start_row+4}'].font = header_font
    
    def _create_conclusion_section(self, sheet: Worksheet):
        """Создает заключительный текст"""
        start_row = len(sheet['B']) + 6
        
        sheet.merge_cells(f'A{start_row}:F{start_row}')
        sheet[f'A{start_row}'] = (
            "Вышеперечисленные услуги выполнены полностью и в срок. "
            "Заказчик претензий по объему, качеству и срокам оказания услуг не имеет."
        )
        sheet[f'A{start_row}'].alignment = DocumentStyles.get_wrap_text_alignment()
        sheet[f'A{start_row}'].font = DocumentStyles.get_normal_font()
        sheet.row_dimensions[start_row].height = 28
    
    def _create_signatures_section(self, sheet: Worksheet):
        """Создает секцию подписей"""
        start_row = len(sheet['B']) + 2
        header_font = DocumentStyles.get_header_font()
        
        # Заголовки подписей
        sheet[f'A{start_row}'] = "ИСПОЛНИТЕЛЬ"
        sheet[f'A{start_row}'].font = header_font
        
        sheet[f'E{start_row}'] = "ЗАКАЗЧИК"
        sheet[f'E{start_row}'].font = header_font
        
        # Названия организаций
        sheet[f'A{start_row+1}'] = self.company.name
        sheet[f'A{start_row+1}'].font = header_font
        
        sheet[f'E{start_row+1}'] = self.customer.name
        sheet[f'E{start_row+1}'].font = header_font
        
        # Подписи руководителей
        sheet[f'A{start_row+3}'] = "Руководитель"
        sheet[f'A{start_row+3}'].font = header_font
        
        try:
            sheet.unmerge_cells(f'C{start_row+3}:D{start_row+3}')
        except:
            pass
        sheet.merge_cells(f'C{start_row+3}:D{start_row+3}')
        sheet[f'C{start_row+3}'] = "_______________"
        sheet[f'C{start_row+3}'].font = header_font
    
    def _amount_to_words(self, amount: float) -> str:
        """Преобразует сумму в слова"""
        try:
            from num2words import num2words
            amount_str = num2words(int(amount), lang='ru')
            amount_str = amount_str[0].upper() + amount_str[1:]
            return f"{amount_str} рублей 00 копеек. Без НДС."
        except:
            return f"{amount:,.2f} рублей 00 копеек. Без НДС."